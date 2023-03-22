from exportcomments import ExportComments, base
from exportcomments.settings import DEFAULT_BASE_URL
import pandas as pd
from pandas.io.excel._openpyxl import OpenpyxlReader
from pandas._typing import FilePathOrBuffer, Scalar
import numpy as np
import requests, os, time, json
############ Importing configuration from config.json  ############
path=os.path.dirname(__file__)
with open(os.path.join(path,'config.json'), 'r') as f:
    config = json.load(f)
if not os.path.exists(os.path.join(path,'temp')): os.makedirs(os.path.join(path,'temp'))
############ Please only edit the following three lines ############
input_=config['input'] #Input file
output_=config['output'] #Output file
token=config['token'] #ExportComments API Access Token
#################################
ex=ExportComments(token)
version=base.version
mod=base.ModelEndpointSet(token=token, base_url=DEFAULT_BASE_URL)
#Loading Log data if the logfile exists
logfile=os.path.join(path,config['log']) #The log file
if not os.path.isfile(logfile): IDs,GUIDs,STATUS=([] for i in range(3))
else:
    log=pd.read_excel(logfile)
    IDs=log['ID'].values.tolist()
    GUIDs=log['GUID'].values.tolist()
    STATUS=log['STATUS'].values.tolist()
df=pd.read_excel(input_)
#GetPostID (FROM url) Getting the Facebook Post ID from the post URL
def GetPostID(url):
    lista=[str(x) for x in range(10)]
    url=url[url.find('m')+2:]+"azezea"
    new,new2,new1=("",)*3
    for i in range(len(url)):
        if url[i] in lista:new+=url[i]
        else: break
    for i in range(len(url)):
        if url[i] in ["/"] and url[i+1] in lista:
            new2=url[i+1:]
            for j in range(len(new2)) :
                if new2[j] in lista: new1+=new2[j]
    return str(new1)
#Just to check the internet connection if everything is okay
def ReTry(data,type):
    while True:
        if requests.get("https://www.google.com").status_code==200:
            if type=='GUID': return ex.exports.check(guid=data)
            elif type=='URL': return ex.exports.create(url=data, replies='false')
            elif type=='GET': return requests.get(data)
        time.sleep(2)
#Save XLSX files
def SaveFile(file_path, URL):
    resp=ReTry(URL,'GET')
    with open(file_path, 'wb') as output: output.write(resp.content)
#Dawnloading ready files
def Download(log):
    while 'progress' in list(log['STATUS']):
        for i in range (len(log)):
            file_path=os.path.join(path,'temp',str(log['ID'][i])+".xlsx")
            if not os.path.isfile(file_path):
                r=ReTry(log['GUID'][i],'GUID').body #Getting information on the status of the request
                new=r['data'][0]['status']
                if new!='progress':
                    log['STATUS'][i]=new #UPDATE the status of eatch Request
                    if new=='done': SaveFile(file_path,"https://exportcomments.com"+r['data'][0]['downloadUrl'])
            time.sleep(20)
    return log
#Configuration of *openpyxl*
def _convert_cell(self, cell, convert_float: bool) -> Scalar:
    from openpyxl.cell.cell import TYPE_BOOL, TYPE_ERROR, TYPE_NUMERIC
    if cell.hyperlink and cell.hyperlink.target: return cell.hyperlink.target
    elif cell.is_date: return cell.value
    elif cell.data_type == TYPE_ERROR: return np.nan
    elif cell.data_type == TYPE_BOOL: return bool(cell.value)
    elif cell.value is None: return ""
    elif cell.data_type == TYPE_NUMERIC:
        if convert_float:
            if int(cell.value) == cell.value: return int(cell.value)
        else: return float(cell.value)
    return cell.value
def load_workbook(self, filepath_or_buffer: FilePathOrBuffer):
    from openpyxl import load_workbook
    return load_workbook(filepath_or_buffer, read_only=False, data_only=True, keep_links=False)
def Merge():
    COM,ID_COM,Post_ID=([] for i in range(3))
    for file in os.listdir(os.path.join(path,'temp')):
        dx = pd.read_excel(os.path.join(path,'temp',file))
        for i in range(5,len(dx)-1):
            COM.append(dx[dx.columns[7]][i])
            ID_COM.append( dx[dx.columns[8]][i].split('/')[-1])
            Post_ID.append(file.split('.')[0])
    dx=pd.DataFrame({'ID_POST': Post_ID, 'ID_COM': ID_COM, 'COM': COM })
    dx.to_excel(os.path.join(path,output_))
############ Code body ############
OpenpyxlReader._convert_cell = _convert_cell
OpenpyxlReader.load_workbook = load_workbook
for i in range(len(df)):
    url=df["URL"][i]
    ID=GetPostID(url)
    if df['Comments'][i] > 0 and not os.path.isfile(os.path.join(path,'temp',str(ID)+".xlsx")):
        #Check the possibility to upload a post URL; Maximum parallel requests=4
        while True:
            rep=ReTry(url,'URL').body
            if len(rep['data'])==4:time.sleep(20)
            else:
                IDs.append(ID)
                GUIDs.append(rep['data']['guid'])
                STATUS.append('progress')
                break
log=pd.DataFrame({'ID': IDs, 'GUID': GUIDs, 'STATUS': STATUS})
log=Download(log)
log.to_excel(logfile) 
Merge() #Merge downloaded files
